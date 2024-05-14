import sqlite3
from openpyxl import Workbook

def create_report_from_database(database_file, output_file):
  
    connection = sqlite3.connect("hotel_test.db")
    cursor = connection.cursor()

    # запрос для rooms
    cursor.execute("SELECT * FROM rooms")
    rooms_rows = cursor.fetchall()

    # запрос для guests
    cursor.execute("SELECT * FROM guests")
    guests_rows = cursor.fetchall()

    # запрос для bookings
    cursor.execute("SELECT * FROM bookings")
    bookings_rows = cursor.fetchall()

    # создаем 3 листа для отображения данных
    wb = Workbook()
    ws_rooms = wb.create_sheet(title="Rooms")
    ws_guests = wb.create_sheet(title="Guests")
    ws_bookings = wb.create_sheet(title="Bookings")

    # дескрипшн нужен для вывода наименования столбцов.
    ws_rooms.append([description[0] for description in cursor.description])
    for row in rooms_rows:
        ws_rooms.append(row)

    ws_guests.append([description[0] for description in cursor.description])
    for row in guests_rows:
        ws_guests.append(row)

    ws_bookings.append([description[0] for description in cursor.description])
    for row in bookings_rows:
        ws_bookings.append(row)

    wb.save(output_file)


    connection.close()

# создание отчета из указанной бд, с указанием наименования отчета
create_report_from_database("hotel_test.db", "hotel_test_report.xlsx")
